from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from .driver import create_driver
import openpyxl
from datetime import date

class Scrappy:
    def __init__(self, browser:str, url:str) -> None:      
        # Creating list of names
        self.companies_names = []
        # Creating list of values
        self.companies_values = []
        # Validating browser
        browser = browser.strip().lower()
        if browser in ["chrome", "edge", "firefox"]:
            self.driver = create_driver(browser)
            self.driver.get(url)
        else:
            raise ValueError("Invalid Browser")

    # Getting names
    def get_names_of_companies(self):
        # Adding wait to loading page
        wait = WebDriverWait(self.driver, 30)

        # Getting name of companies
        names_companies = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "mn_merchName")))

        # Adding interator
        for name in names_companies:
            name = name.text.strip()
            if name:
                self.companies_names.append(name)
        return self.companies_names

    # Getting values
    def get_values_of_companies(self):
        # Adding wait to loading page
        wait = WebDriverWait(self.driver, 50)

        # Getting value of companies
        values_companies = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "mn_rebate")))

        # Adding interator
        for value_companie in values_companies:
            value = value_companie.text.strip()[5:]
            if value:
                value = value.split("\n")[0]
                if " " in value:
                    values_split = value.split(" ")
                    if len(values_split) > 2:
                        value_list = [values_split[-1]]
                        value_split = values_split[-2]
                        value_list.insert(0, value_split)
                        value = " ".join(value_list)
                        if "/$" not in value:
                            value += "/$"
                self.companies_values.append(value)
        return self.companies_values

    def generate_plan(self):
        # Getting name and values
        self.get_names_of_companies()
        self.get_values_of_companies()

        # Number of cell when start data
        index = 2
        # Instancing workbook
        workbook = openpyxl.Workbook()
        # Initializating worbook
        sheet = workbook.active
        # Configurated title
        sheet.title = "Companies"
        # Names of sheet
        sheet["A1"] = "Names"
        sheet["B1"] = "Values"
        for name, value in zip(self.companies_names, self.companies_values):
            sheet.cell(column=1, row=index, value=name)
            sheet.cell(column=2, row=index, value=value)
            sheet.cell(column=3, row=index value=date.date())
            index += 1
        workbook.save("companies_infos.xlsx")

if __name__ == "__main__":
    scrap = Scrappy("chrome", "https://www.aadvantageeshopping.com/b____.htm")
    scrap.generate_plan()